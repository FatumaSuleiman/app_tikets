
from django.shortcuts import render
from firebase_connection import db
from rest_framework import status
from rest_framework.views import APIView
from rest_framework.response import Response
from datetime import datetime
import math
import random
from django.views.decorators.http import require_http_methods
from openpyxl import Workbook
from django.http import HttpResponse
from ticket_app.settings import env
from transactionsaver.views.users_views import FirebaseAuthentication



def institution_events(request, institution_id, phone_number):
    """
        This method is for getting institution active events 
        and render them on institution events page with its 
        institution identifier bundled to the institution
    """
    try:
        responses = db.collection('institutions').document(institution_id).collection('events').get()
        context = []
        for doc in responses:
            # print(doc.id)
            if doc.to_dict()['active']:
                docId =  doc.id 
                combinedData = {
                    'id':docId,
                    'data':doc.to_dict()
                }
                context.append(combinedData)
        return render(request, 'institutions_events.html', {'events':context, 'institutionId':institution_id, 'phone_number':phone_number})
    except Exception as e:
        print(e)





def ticket_category(request, institution_id, event_id, phone_number):
    """
    This method is for getting ticket categories for a specific
    events.
    """
    try:
        groupUsedTickets = 1
        responses = db.collection(u'institutions').document(institution_id).collection('events').document(event_id).collection('ticket_category').order_by(
    u'amount').get()
        eventData = db.collection(u'institutions').document(institution_id).collection('events').document(event_id).get()    
        context = []
        for doc in responses:

            """
            This condition is for counting number of individual tickets with 
             group id. then calculate the quantity of the group remains.
             by dividing the number with the number of people who can seat
             on a group (table) .
            """
            if doc.to_dict()['is_group']:
                print(doc.to_dict()['category_id'])
                individualTicketWithGroupId = db.collection(u'institutions').document(institution_id).collection('events').document(event_id).collection('tickets').where(u'ticketCategoryId','==', doc.to_dict()['category_id']).get() 
                groupUsedTickets = len(individualTicketWithGroupId) / int(doc.to_dict()['group_quantity'])
            
            """
            request below is for counting the number of tickets used per category.
            """

            res = db.collection(u'institutions').document(institution_id).collection('events').document(event_id).collection('tickets').where(u'ticketCategoryId','==', doc.to_dict()['category_id']).get()          
            combined = {
                'remainingIndividualTickets':int(doc.to_dict()['quantity']) - len(res),
                'data':doc.to_dict(),
                'remainingGroupTickets':int(doc.to_dict()['quantity']) - round(groupUsedTickets)
            }
            context.append(combined)
        return render(request, 'event_ticket_category.html', {'ticket_categories':context, 'eventMomo':eventData.to_dict()['MomoCode'],'event_id':event_id, 'institution_id':institution_id, 'eventDevice':eventData.to_dict()['eventDevice'], 'phone_number':phone_number})
    except Exception as e:
        print(e)


def institution_event_stats(request, institution_id):
    """
        This method is for getting institution active events 
        and render them on institution events page with its 
        institution identifier bundled to the institution
    """
    try:
        responses = db.collection('institutions').document(institution_id).collection('events').get()
        context = []
        for doc in responses:
            # print(doc.id)
            # if doc.to_dict()['active']:
                docId =  doc.id
                combinedData = {
                    'id':docId,
                    'data':doc.to_dict()
                }
                context.append(combinedData)
        return render(request, 'institutions_event_stats.html', {'events':context, 'institutionId':institution_id})
    except Exception as e:
        print(e)



class SaveCardTransaction(APIView):
    """
        This Method is for checking the saving the card transaction before processing card payment,
        after card payment we use the transaction number saved in cardTransactions as payment 
        transaction number and save orginal transaction to transactions of events collection.
    """
    authentication_classes=(FirebaseAuthentication,)
    def post(self, *args, **kwargs):
        try:
            min_n = 15245189
            max_n = 987654321987
            min_n = math.ceil(min_n)
            max_n = math.floor(max_n)
            paymentNumber=math.floor(random.randint(15, 10000)*(max_n - min_n + 1)+min_n)

            institution_id = self.kwargs['institution_id']
            event_id = self.kwargs['event_id']

            transaction_object = {
                'paymentNumber':paymentNumber,
                'amount':self.request.data['amount'],
                'payerEmail':self.request.data['payerEmail'],
                'ticketCategory':self.request.data['ticketCategory'],
                'ticketQuantity':self.request.data['ticketQuantity'],
                'paymentTime':datetime.now(),
            }
            db.collection(u'institutions').document(institution_id).collection('events').document(event_id).collection('cardTransactions').add(transaction_object)
            return Response(data=transaction_object, status=status.HTTP_200_OK)

        except Exception as e:
            print(e)
            return Response(data="Internal Server error", status=status.HTTP_500_INTERNAL_SERVER_ERROR)




@require_http_methods(["GET"])
def success(request):
    return render(request, 'success.html')




def institution_event_free_tickets(request, institution_id):
    """
        This method is for getting institution active events 
        and render them on institution events page with its 
        institution identifier bundled to the institution
    """
    try:
        responses = db.collection('institutions').document(institution_id).collection('events').get()
        context = []
        for doc in responses:
            # print(doc.id)
            # if doc.to_dict()['active']:
                docId =  doc.id
                free_ticket_category_object=[]
                free_ticket_category = db.collection('institutions').document(institution_id).collection('events').document(docId).collection('ticket_category').where('is_free','==',True).get()
                if len(free_ticket_category)>0:
                    free_ticket_category_object=free_ticket_category[0].to_dict()

                all_free_categories=[]
                for c in free_ticket_category:
                    all={
                        'id':c.id,
                        'data':c.to_dict()
                    }
                    
                    all_free_categories.append(all)
                

                combinedData = {
                    'id':docId,
                    'data':doc.to_dict(),
                    'free_ticket_category':all_free_categories
                }
                
                context.append(combinedData)
        return render(request, 'institutions_event_freetickets.html', {'events':context, 'institutionId':institution_id})
    except Exception as e:
        print(e)



class GetTicketCategories(APIView):
    """
    This Method is for fetching ticket categories and be used
    before merging the tickets
    """
    authentication_classes=(FirebaseAuthentication,)
    def post(self, *args, **kwargs):
        try:
            tickectCategory = []
            instId = self.request.data['instId']
            eventId = self.request.data['eventId']
            documents = db.collection(u'institutions').document(instId).collection('events').document(eventId).collection('ticket_category').where(u'is_free', '==', False).get()
            if documents:
                for doc in documents:
                    docId = doc.id
                    combined_data = {
                        'id': docId,
                        'data': doc.to_dict()
                    }
                    tickectCategory.append(combined_data)
                return Response(data=tickectCategory, status=status.HTTP_200_OK)
            else:
                return Response(data="Ticket Has not categories", status=status.HTTP_204_NO_CONTENT)

        except Exception as e:
            print(e)
            return Response(data="Internal Server error", status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class MergeTickets(APIView):
    """
    This Method merging a ticket with another ticket, and validate one
    of them so that it can be used and also 
    """
    authentication_classes=(FirebaseAuthentication,)
    def post(self, *args, **kwargs):
        try:
            instId = self.request.data['instId']
            eventId = self.request.data['eventId']
            ticketCategoryId = self.request.data['ticketCategoryId']
            existingTicketNumber = self.request.data['existingTicketNumber']
            newTicketNumber = self.request.data['newTicketNumber']

            # get the ticketCategory object (in order to get ticket amount)
            ticketCategory = db.collection(u'institutions').document(instId).collection(
                'events').document(eventId).collection('ticket_category').document(ticketCategoryId).get()

            # this is ticket amount for selected category.
            ticketAmount = ticketCategory.to_dict()['amount']
  

            # get tickets objects
            existingTicketObj = db.collection(u'institutions').document(instId).collection('events').document(
                eventId).collection('tickets').where(u'ticketNumber', '==', existingTicketNumber).get()

            newTicketObj = db.collection(u'institutions').document(instId).collection('events').document(
                eventId).collection('tickets').where(u'ticketNumber', '==', newTicketNumber).get()

            if newTicketObj:
                # get both tickets amount and , sum of their amount
                existingTicketAmount = existingTicketObj[0].to_dict()['amount']
                newTicketAmount = newTicketObj[0].to_dict()['amount']
                ticketsAmountSum = int(existingTicketAmount) + int(newTicketAmount)

                if int(ticketAmount) <= ticketsAmountSum:
                    # here you can continue merging tickets
                    existingTicketObj[0].reference.update({u'valid': True, u'comment': "merged with ticket {}".format(
                        newTicketNumber), 'mergedAt': datetime.now()})
                    newTicketObj[0].reference.update({u'valid': False, u'comment': "merged with ticket {}".format(
                        existingTicketNumber), 'mergedAt': datetime.now()})
                    return Response(data='Tickets Merged Successfully', status=status.HTTP_200_OK)
                else:
                    # still amount is not enough according to selected ticket category
                    return Response(data='Not enough amount to Merge tickets.', status=status.HTTP_412_PRECONDITION_FAILED)
            else:
                return Response(data='Tickets Not Found, try again', status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response(data="Internal Server error", status=status.HTTP_500_INTERNAL_SERVER_ERROR)



class GetTicketsFromCategory(APIView):
    """
    This Method is for fetching ticket by categories 
    """
    authentication_classes=(FirebaseAuthentication,)
    def get(self, *args, **kwargs):
        try:
            tickets = []
            instId = kwargs['instId']
            categoryId = kwargs['categoryId']
            eventId = kwargs['eventId']


            # # get institution identifier
            inst_doc = db.collection(u'institutions').document(instId).get()
            ins_obj = inst_doc.to_dict()
            institution_identifier = str(ins_obj['identifier'])

            # # get event identifier
            event_doc = db.collection(u'institutions').document(instId).collection(u'events').document(eventId).get()
            event_obj = event_doc.to_dict()
            event_identifier = str(event_obj['eventID'])
            event_name = str(event_obj['name'])

            # return Response(data=str(institution_identifier) + " " +event_identifier, status=status.HTTP_200_OK)


            documents = db.collection(u'institutions').document(instId).collection('events').document(eventId).collection('tickets').where(u'ticketCategoryId', '==', categoryId).get()
            if documents and event_identifier and institution_identifier and event_name:

                filename="tickets_for_"+event_name+".xlsx"
                response = HttpResponse(content_type='application/ms-excel')
                #response['Content-Disposition'] = 'attachment; filename="payroll.xls"'
                response['Content-Disposition'] = 'attachment; filename="'+filename+'"'
            
                wb = Workbook()

                # grab the active worksheet
                ws = wb.active
                ws=wb.create_sheet('workshift',0)

                # Data can be assigned directly to cells
                ws['A1'] = 'Name'
                ws['B1'] = 'Amount'
                ws['C1'] = 'Payment Number'
                ws['D1'] = 'Payment Time'
                ws['E1'] = 'Phone'
                ws['F1'] = 'Ticket Number'
                ws['G1'] = 'Valid'
                ws['H1'] = 'Validated'
                ws['I1'] = 'Ticket Url'

                #Loop

                for doc in documents:
                    data = doc.to_dict()
                    ws.append([data['name'], data['amount'], data['paymentNumber'], data['paymentTime'], data['phone'], data['ticketNumber'], data['valid'],
                              data['validated'],env("NOKANDA_TICKET_APP_URL")+institution_identifier+'/'+event_identifier+'/'+data['ticketNumber']+''])


                    tickets.append(doc.to_dict())
                # return Response(data=tickets, status=status.HTTP_200_OK)
                wb.save(response) 
                return response
            else:
                return Response(data="Ticket Does not exist", status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            print(e)
            return Response(data="Internal Server error", status=status.HTTP_500_INTERNAL_SERVER_ERROR)
