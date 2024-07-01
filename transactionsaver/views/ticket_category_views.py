from firebase_connection import db
from rest_framework import status
import threading
from rest_framework.views import APIView
from rest_framework.response import Response
from transactionsaver import serializers
from rest_framework.generics import GenericAPIView
from .transaction_views import get_ticket_number,get_phone_fromReason
from transactionsaver.views.users_views import FirebaseAuthentication
from .transaction_views import createInstitutionEventTransaction,createInstitutionEventTickets,is_email,send_message,send_ticket_in_email,createInstitutionTickets,SendSMSAndSaveData_V2
from django.http import JsonResponse
class SendSMSAndSaveData1(threading.Thread):
    def __init__(self,tickets, transObj, tr_type):
        self.tickets = tickets
        self.transObj = transObj
        self.tr_type=tr_type
        threading.Thread.__init__(self)

    def run(self):
        send_to_firebase_and_messages(self.tickets,self.transObj,self.tr_type)

    def join(self, timeout=None):
        """ Stopping the thread. """
        self._stopevent.set()
        threading.Thread.join(self, timeout)



def send_to_firebase_and_messages(tickets,trans_obj,tr_type):
    """
    this is the method to save the transaction and ticket of
    a given institution (entity) to firebase and then
    send the qr message to the corresponding user.
    """

    institution_id = trans_obj['instId']
    eventId=trans_obj['eventId']
    institution_identifier=0
    eventID=0
    eventName=''
    # find event of the institution that is using the device for paying
    event_doc = db.collection(u'institutions').document(institution_id).collection('events').where('eventId','==',eventId).get()
    
    # Find identifier of institution
    inst_doc=db.collection(u'institutions').document(institution_id).get()
    ins_obj=inst_doc.to_dict()
    institution_identifier=ins_obj['identifier']
    #createInstitutionTransaction(trans_obj=trans_obj, inst_id=institution_id)
    create_ticket=True
    if len(event_doc)>0:
        create_ticket=createInstitutionEventTransaction(trans_obj=trans_obj, inst_id=institution_id,event_id=event_doc[0].id)
        eventID=event_doc[0].to_dict()['eventID']
        eventName=event_doc[0].to_dict()['name']
    if create_ticket==False:
        print("We couldn't create ticket due to the transaction number that is alread exist"+trans_obj['paymentNumber'])
    reason_phone=''
    if is_email(trans_obj["reason"])==False:    
        reason_phone = get_phone_fromReason(trans_obj["reason"])
    if tr_type == "INSTITUTION" and create_ticket:
        for ticket in tickets:
            check_data = db.collection("tickets").where('paymentNumber', '==', ticket["paymentNumber"]).get()
            #if len(check_data) == 0:
            db.collection("tickets").add(ticket)
            if len(tickets) > 1:
                if ticket["phone"]!="N/A":
                    send_message(ticket["phone"], ticket["ticketNumber"], inst_identifier=institution_identifier,event_id=eventID,event_name=eventName)
            else:
                if ticket["phone"]!="N/A":
                    send_message(reason_phone, ticket["ticketNumber"],inst_identifier=institution_identifier,event_id=eventID,event_name=eventName)
    elif tr_type == "INDIVIDUAL" and create_ticket:
        #check_data = db.collection("tickets").where('paymentNumber', '==', trans_obj["paymentNumber"]).get()
        #if len(check_data) == 0: 
        for ticket in tickets:
                #createInstitutionTickets(ticket=ticket, inst_id=institution_id)
                if len(event_doc)>0:
                    createInstitutionEventTickets(ticket=ticket, inst_id=institution_id,event_id=event_doc[0].id)
                    if len(tickets) > 1:
                        if ticket["phone"]!="N/A":
                            send_message(ticket["phone"], ticket["ticketNumber"],inst_identifier=institution_identifier,event_id=eventID,event_name=eventName)
                        
                        if len(ticket["reason"])>1:
                            send_ticket_in_email(ticket["reason"], ticket["ticketNumber"],inst_identifier=institution_identifier,event_id=eventID,event_name=eventName)
                     
                    else:
                        if len(reason_phone)>9:
                            if ticket["phone"]!="N/A":
                                send_message(reason_phone, ticket["ticketNumber"],inst_identifier=institution_identifier,event_id=eventID,event_name=eventName)
                            
                            if len(ticket["reason"])>1:
                                send_ticket_in_email(ticket["reason"], ticket["ticketNumber"],inst_identifier=institution_identifier,event_id=eventID,event_name=eventName)
                     
                        else:
                            if ticket["phone"]!="N/A":
                                send_message(ticket["phone"], ticket["ticketNumber"],inst_identifier=institution_identifier,event_id=eventID,event_name=eventName)
                            
                            if len(ticket["reason"])>1:
                                send_ticket_in_email(ticket["reason"], ticket["ticketNumber"],inst_identifier=institution_identifier,event_id=eventID,event_name=eventName)
                     
    else:
        for ticket in tickets:
            if create_ticket:
                createInstitutionTickets(ticket=ticket, inst_id=institution_id)
            # db.collection("tickets").add(ticket)
            if len(tickets) > 1:
                if ticket["phone"]!="N/A" and create_ticket:
                 send_message(ticket["phone"], ticket["ticketNumber"],inst_identifier=institution_identifier,event_id=eventID,event_name=eventName)
            else:
                if ticket["phone"]!="N/A" and create_ticket:
                    send_message(reason_phone, ticket["ticketNumber"],inst_identifier=institution_identifier,event_id=eventID,event_name=eventName)

def send_to_firebase_and_message2(tickets,trans_obj,tr_type,number_of_allowed_ticket):

    reason_phone = get_phone_fromReason(trans_obj["reason"])
    if tr_type == "INSTITUTION":
        for ticket in tickets:
            check_data = db.collection("tickets").where('paymentNumber', '==', ticket["paymentNumber"]).get()
            if number_of_allowed_ticket> len(check_data) :
                db.collection("tickets").add(ticket)
                if len(tickets) > 1:
                    send_message(ticket["phone"], ticket["ticketNumber"])
                else:
                    send_message(trans_obj["phone"], ticket["ticketNumber"])
                        # Send message to reason phone
                    #send_message(reason_phone, ticket["ticketNumber"])
    elif tr_type == "INDIVIDUAL":
        check_data = db.collection("tickets").where('paymentNumber', '==', trans_obj["paymentNumber"]).get()
        if len(check_data) == 0:
            for ticket in tickets:
                db.collection("tickets").add(ticket)
                if len(tickets) > 1:
                    send_message(ticket["phone"], ticket["ticketNumber"])
                else:
                    send_message(trans_obj["phone"], ticket["ticketNumber"])
                    # Send message to reason phone
                    send_message(reason_phone, ticket["ticketNumber"])
    else:
        for ticket in tickets:
            db.collection("tickets").add(ticket)
            if len(tickets) > 1:
                send_message(ticket["phone"], ticket["ticketNumber"])
            else:
                send_message(trans_obj["phone"], ticket["ticketNumber"])
                # Send message to reason phone
                send_message(reason_phone, ticket["ticketNumber"])
class AddTicketCategory(GenericAPIView):
    serializer_class = serializers.TicketCategorySerializer
    authentication_classes = (FirebaseAuthentication,)

    def post(self, request, institution_id, event_id):
        category_data = {
            'category_id': self.request.data['category_id'],
            'name': self.request.data['name'],
            'amount': self.request.data['amount'],
            'quantity': self.request.data['quantity'],
            'group_quantity': self.request.data['group_quantity'],
            'is_free': self.request.data['is_free'],
            'is_group': self.request.data['is_group']
        }
        cat_id = self.request.data['category_id']
        event_ref = db.collection("institutions").document(institution_id).collection("events")
        docs = event_ref.document(event_id).collection("ticket_category").get()

        if len(docs) == 0:
            event_ref.document(event_id).collection("ticket_category").add(category_data)
            return JsonResponse({'data': category_data}, status=status.HTTP_201_CREATED)
        else:
            check = False
            for doc in docs:
                data = doc.to_dict()
                if cat_id == data['category_id']:
                    check = True
                    break
            if check:
                return JsonResponse({'data': "category_id already exist"}, status=status.HTTP_400_BAD_REQUEST)
            else:
                event_ref.document(event_id).collection("ticket_category").add(category_data)
                return JsonResponse({'data': category_data}, status=status.HTTP_201_CREATED)
class GetAlleventTicketCategory(APIView):
    # Endpoint for get all ticket_categories  of an event
    authentication_classes=(FirebaseAuthentication,)
    def get(self,request,inst_id,event_id):
        category_list=[]
        results=db.collection("institutions").document(inst_id).collection("events").document(event_id).collection("ticket_category").get()
        for doc in results:
            ticket_category_id=doc.id
            data={
                "ticket_category_id":ticket_category_id,
                "category_data":doc.to_dict()
            }
            category_list.append(data)
        return Response(data={'data':category_list},status=status.HTTP_200_OK)
    
class GetCategoryDetails(GenericAPIView):
    # endpoint for ticket_category details
    serializer_class=serializers.TicketCategorySerializer
    def get(self,request,inst_id,event_id,categ_id):
        result=db.collection("institutions").document(inst_id).collection('events').document(event_id).collection("ticket_category").document(categ_id).get()
        data={
            'categ_id': result.id,
            'category_data':result.to_dict()
        }
        return Response(data=data,status=status.HTTP_200_OK)
class UpdateTicketCategory(GenericAPIView):
    # endpoint for updating TicketCategory
    serializer_class=serializers.TicketCategorySerializer
    authentication_classes=(FirebaseAuthentication,)
    def put(self,request,institution_id,event_id,categ_id):
        categoryId = request.data.get('category_id')
        category_data={
            'category_id':request.data.get('category_id'),
            'name':request.data.get('name'),
            'amount':request.data.get('amount'),
            'quantity': request.data.get('quantity'),
            'group_quantity':request.data.get('group_quantity'),
            'is_free':request.data.get('is_free'),
            'is_group':request.data.get('is_group')
        }

        cate_doc=db.collection("institutions").document(institution_id).collection("events").document(event_id).collection("ticket_category").document(categ_id).get()
        if cate_doc.exists:
            data = cate_doc.to_dict()
            if categoryId == data['category_id']:
                db.collection("institutions").document(institution_id).collection("events").document(event_id).collection("ticket_category").document(categ_id).update(category_data)
                return JsonResponse(data={'data':category_data},status=status.HTTP_201_CREATED)
            else:
                return JsonResponse(data={'data':" categoryId should be unique"})
        else:
             return JsonResponse(data={'data': "Ticket Category not found."})
      

def create_ticket(data):
 
    tickets = []
    # Find event on which a device is linked to and thus get single purchase amount
    institution_id = data['instId']
    event_id=data['eventId']
    event_doc = db.collection("institutions").document(institution_id).collection("events").where('eventID','==',event_id).get()
    singlePurchaseAmount=10000

    if len(event_doc)>0:
        p_object=event_doc[0].to_dict()
        singlePurchaseAmount=int(p_object['SinglePurchaseAmount'])
        print("This is category Id")
        print(data['ticketCategoryId'])
        ticketCategoryId=int(data['ticketCategoryId'])
        #Check the category of ticket being paid and check paid amount are enough
        ticketCategory_doc = db.collection("institutions").document(institution_id).collection("events").document(event_doc[0].id).collection("ticket_category").where('category_id','==',data['ticketCategoryId']).get()
        valid=True    
        
        if len(ticketCategory_doc)>0:
            #There is specific ticket category
            ticketCategory_object=ticketCategory_doc[0].to_dict()
            singlePurchaseAmount=int(ticketCategory_object['amount'])
            
            if ticketCategory_object['is_group']==True:
                print("Yes it's group")
                #Ticket category is for group
                if singlePurchaseAmount>int(data['amount']):
                    valid=False
                group_quantity=0
                if ticketCategory_object['group_quantity']:
                    group_quantity=int(ticketCategory_object['group_quantity'])
                number_of_ticket=group_quantity
                

                if p_object['groupnumber_assignation_required']==True:
                    #Group/table assignation for the event is required
                    how_many_per_group=p_object['how_many_per_group']
                    unfinished_group=p_object['unfinished_group']
                    unfinished_group_remaining_members=p_object['unfinished_group_remaining_members']
                    current_group_number=p_object['current_group_number']
                    ticket_group_number=0
                    
                    new_unfinished_group=unfinished_group
                    new_unfinished_group_remaining_members=unfinished_group_remaining_members
                    new_current_group_number=current_group_number 
                    print('how_many_per_group')
                    print(how_many_per_group)
                    print('group_quantity')
                    print(group_quantity)
                    if group_quantity==how_many_per_group:
                        # The quantity of member of the group is equal 
                        # to number of people that should be in a group in this event
                        if group_quantity>unfinished_group_remaining_members:
                           ticket_group_number=current_group_number+1
                           new_current_group_number=ticket_group_number
                        elif group_quantity<unfinished_group_remaining_members:
                            ticket_group_number=unfinished_group
                            new_unfinished_group_remaining_members=new_unfinished_group_remaining_members-group_quantity
                        else:
                            new_current_group_number=new_current_group_number+1
                            ticket_group_number=unfinished_group
                            new_unfinished_group=new_current_group_number
                            new_unfinished_group_remaining_members=how_many_per_group

                        
                    elif group_quantity>how_many_per_group:
                        # The quantity of member of the group is greater 
                        # to number of people that should be in a group in this event
                        pass
                    else:
                        # The quantity of member of the group is less 
                        # to number of people that should be in a group in this event
                        if group_quantity>unfinished_group_remaining_members:
                           ticket_group_number=current_group_number+1
                           new_current_group_number=ticket_group_number
                        elif group_quantity<unfinished_group_remaining_members:
                            ticket_group_number=unfinished_group
                            new_unfinished_group_remaining_members=new_unfinished_group_remaining_members-group_quantity
                        else:
                            ticket_group_number=unfinished_group
                            new_unfinished_group=new_current_group_number
                            new_unfinished_group_remaining_members=how_many_per_group
                    
                    # Creating ticket

                    # If it's free ticket
                    if ticketCategory_object['is_free']==True:
                        number_of_ticket=int(float(data['quantity']))

                    if number_of_ticket == 1:
                        # if len(ticketCategory_doc)>0:
                        #     ticketCategory_object=ticketCategory_doc.to_dict()
                        #     if int(ticketCategory_object['amount'])>int(data['amount']):
                        #         valid=False
                        

                        ticket_obj = {
                            "amount": data['amount'],
                            "name": data["name"],
                            "paymentNumber": data["paymentNumber"],
                            "paymentTime": data["paymentTime"],
                            "phone": data["phone"],
                            "reason": data["reason"],
                            "validatedAt": "",
                            'ticketNumber': str(get_ticket_number()),
                            'singlePurchaseAmount': singlePurchaseAmount,
                            'validated':False,
                            'validatedBy':'',
                            'ticketCategoryId':data['ticketCategoryId'],
                            'valid':valid,
                            'ticket_group_number':ticket_group_number
                        }
                        tickets.append(ticket_obj)
                    else:
                        for i in range(0, number_of_ticket):
                           # if int(data['amount'])/(singlePurchaseAmount*(i+1))<1:
                            #    valid=False

                            ticket_obj = {
                                "amount": singlePurchaseAmount,
                                "name": data["name"],
                                "paymentNumber": data["paymentNumber"],
                                "paymentTime": data["paymentTime"],
                                "phone": data["phone"],
                                #"reason": get_phone_fromReason(data["reason"]),
                                "reason": data["reason"],
                                "validatedAt": "",
                                'ticketNumber': str(get_ticket_number()),
                                'singlePurchaseAmount':singlePurchaseAmount,
                                'validated':False,
                                'validatedBy':'',
                                'ticketCategoryId':data['ticketCategoryId'],
                                'valid':valid,
                                'ticket_group_number':ticket_group_number
                            }
                            tickets.append(ticket_obj)


                    db.collection(u'institutions').document(institution_id).collection('events').document(event_doc[0].id).update({'current_group_number':new_current_group_number,'unfinished_group':new_unfinished_group,'unfinished_group_remaining_members':new_unfinished_group_remaining_members})
                else:

                    #Group/table assignation for the event is not required
                    # if it's for free ticket
                    if ticketCategory_object['is_free']==True:
                        number_of_ticket=int(float(data['quantity']))

                    if number_of_ticket == 1:
                        # if len(ticketCategory_doc)>0:
                        #     ticketCategory_object=ticketCategory_doc.to_dict()
                        #     if int(ticketCategory_object['amount'])>int(data['amount']):
                        #         valid=False
                        

                        ticket_obj = {
                            "amount": data['amount'],
                            "name": data["name"],
                            "paymentNumber": data["paymentNumber"],
                            "paymentTime": data["paymentTime"],
                            "phone": data["phone"],
                            "reason": data["reason"],
                            "validatedAt": "",
                            'ticketNumber': str(get_ticket_number()),
                            'singlePurchaseAmount': singlePurchaseAmount,
                            'validated':False,
                            'validatedBy':'',
                            'ticketCategoryId':data['ticketCategoryId'],
                            'valid':valid
                        }
                        tickets.append(ticket_obj)
                    else:
                        for i in range(0, number_of_ticket):
                           # if int(data['amount'])/(singlePurchaseAmount*(i+1))<1:
                            #    valid=False

                            ticket_obj = {
                                "amount": singlePurchaseAmount,
                                "name": data["name"],
                                "paymentNumber": data["paymentNumber"],
                                "paymentTime": data["paymentTime"],
                                "phone": data["phone"],
                                "reason": get_phone_fromReason(data["reason"]),
                                "validatedAt": "",
                                'ticketNumber': str(get_ticket_number()),
                                'singlePurchaseAmount':singlePurchaseAmount,
                                'validated':False,
                                'validatedBy':'',
                                'ticketCategoryId':data['ticketCategoryId'],
                                'valid':valid
                            }
                            tickets.append(ticket_obj)

                
            else:

                # Ticket category is not for group/table
                Amount=int(data['amount'])
                SinglePurchaseAmount=int(singlePurchaseAmount)
                if SinglePurchaseAmount <= 1 or SinglePurchaseAmount > Amount:
                    number_of_ticket = 1
                else:
                    number_of_ticket = Amount / SinglePurchaseAmount

                if number_of_ticket<1:
                    number_of_ticket=1 

                # if it's for free ticket
                if ticketCategory_object['is_free']==True:
                        number_of_ticket=int(float(data['quantity']))

                if number_of_ticket == 1:
                    # if len(ticketCategory_doc)>0:
                    #     ticketCategory_object=ticketCategory_doc.to_dict()
                    #     if int(ticketCategory_object['amount'])>int(data['amount']):
                    #         valid=False
                    if singlePurchaseAmount>int(data['amount']):
                        valid=False

                    ticket_obj = {
                        "amount": data['amount'],
                        "name": data["name"],
                        "paymentNumber": data["paymentNumber"],
                        "paymentTime": data["paymentTime"],
                        "phone": data["phone"],
                        "reason": data["reason"],
                        "validatedAt": "",
                        'ticketNumber': str(get_ticket_number()),
                        'singlePurchaseAmount': singlePurchaseAmount,
                        'validated':False,
                        'validatedBy':'',
                        'ticketCategoryId':data['ticketCategoryId'],
                        'valid':valid
                    }
                    tickets.append(ticket_obj)
                else:
                    for i in range(0, number_of_ticket):
                        if int(data['amount'])/(singlePurchaseAmount*(i+1))<1:
                            if ticketCategory_object['is_free']==False:
                                valid=False

                        ticket_obj = {
                            "amount": singlePurchaseAmount,
                            "name": data["name"],
                            "paymentNumber": data["paymentNumber"],
                            "paymentTime": data["paymentTime"],
                            "phone": data["phone"],
                            "reason": get_phone_fromReason(data["reason"]),
                            "validatedAt": "",
                            'ticketNumber': str(get_ticket_number()),
                            'singlePurchaseAmount':singlePurchaseAmount,
                            'validated':False,
                            'validatedBy':'',
                            'ticketCategoryId':data['ticketCategoryId'],
                            'valid':valid
                        }
                        tickets.append(ticket_obj)
        else:
            # there is no specific ticket category
            if singlePurchaseAmount <= 1 or singlePurchaseAmount > data['amount']:
                number_of_ticket = 1
            else:
                number_of_ticket = int(data['amount'] / singlePurchaseAmount)

            if number_of_ticket<1:
                number_of_ticket=1 

            if number_of_ticket == 1:
                # if len(ticketCategory_doc)>0:
                #     ticketCategory_object=ticketCategory_doc.to_dict()
                #     if int(ticketCategory_object['amount'])>int(data['amount']):
                #         valid=False
                if singlePurchaseAmount>int(data['amount']):
                    valid=False

                ticket_obj = {
                    "amount": data['amount'],
                    "name": data["name"],
                    "paymentNumber": data["paymentNumber"],
                    "paymentTime": data["paymentTime"],
                    "phone": data["phone"],
                    "reason": data["reason"],
                    "validatedAt": "",
                    'ticketNumber': str(get_ticket_number()),
                    'singlePurchaseAmount': singlePurchaseAmount,
                    'validated':False,
                    'validatedBy':'',
                    'ticketCategoryId':data['ticketCategoryId'],
                    'valid':valid
                }
                tickets.append(ticket_obj)
            else:
                for i in range(0, number_of_ticket):
                    if int(data['amount'])/(singlePurchaseAmount*(i+1))<1:
                        valid=False

                    ticket_obj = {
                        "amount": singlePurchaseAmount,
                        "name": data["name"],
                        "paymentNumber": data["paymentNumber"],
                        "paymentTime": data["paymentTime"],
                        "phone": data["phone"],
                        "reason": get_phone_fromReason(data["reason"]),
                        "validatedAt": "",
                        'ticketNumber': str(get_ticket_number()),
                        'singlePurchaseAmount':singlePurchaseAmount,
                        'validated':False,
                        'validatedBy':'',
                        'ticketCategoryId':data['ticketCategoryId'],
                        'valid':valid
                    }
                    tickets.append(ticket_obj)


    return tickets
    

class SaveTickets(GenericAPIView):
    serializer_class=serializers.TransactionSerializer
    authentication_classes=(FirebaseAuthentication,)
    def post(self,request):
        dat="instId eventId category_id"
        dat1=','.join(str(da) for da in dat)
        seriazed_data = serializers.TransactionSerializer
        referenceNumber=dat1.split()
        referenceNumber1 = request.data['referenceNumber'].split() if isinstance(request.data['referenceNumber'], str) else request.data['referenceNumber']
        if len(referenceNumber) == 3 and len(referenceNumber1) == 3:
            trans_obj={
                'name':request.data['sender'],
                'amount':request.data['amount'],
                'paymentNumber':request.data["transactionNumber"],
                'paymentTime':request.data["transactionTime"],
                'phone':request.data["senderPhone"],
                'instId':referenceNumber1[0],
                 'eventId':referenceNumber1[1],
                'ticketCategoryId':referenceNumber1[2],
                'reason':request.data["reason"],
                'quantity':request.data["quantity"]
            }
            tickets = create_ticket(trans_obj)
            
        # Send data to firebase
            SendSMSAndSaveData_V2(tickets, trans_obj, "INDIVIDUAL").start()
            return Response(data={"transaction":tickets})


        # # Send sms to ticket buyer
        # SendTicketSMSs(tickets, transObj).start()

        else: 
            return Response(data={"data":"referenceNumber is not valid"})

class ValidTicketsForEvent(GenericAPIView):
    serializer_class=serializers.TicketDataSerializer
    authentication_classes=(FirebaseAuthentication,)
    def get(self,request,inst_id,eventId):
        ticket_ref=db.collection("institutions").document(inst_id).collection("events").document(eventId).collection("tickets").where('valid','==',True).get()
           
        return Response(data={"number of tickets":len(ticket_ref)})

class InvalidTicketsForEvent(GenericAPIView):
    serializer_class=serializers.TicketDataSerializer
    authentication_classes=(FirebaseAuthentication,)
    def get(self,request,inst_id,eventId):
        ticket_ref=db.collection("institutions").document(inst_id).collection("events").document(eventId).collection("tickets").where('valid','==',False).get()
           
        return Response(data={"number of tickets":len(ticket_ref)})
import os
from ticket_app import settings
from openpyxl import load_workbook
from datetime import datetime
import openpyxl
from transactionsaver.serializers import BuyTicketsSerializer
class GenerateTicketsFromExcel1(GenericAPIView):
    serializer_class = serializers.FileUploadSerializer

    def post(self, request, *args, **kwargs):
        instId = request.data['instId']
        eventId = request.data['eventId']
        categ_id = request.data['categ_id']
        file = request.data['file']

        tickets = []
        wb = openpyxl.load_workbook(file)
        worksheet = wb["Sheet1"]

        # Query the event document using eventID
        event_docs = db.collection("institutions").document(instId).collection("events").where('eventID', '==', eventId).get()
        if not event_docs:
            return Response(data={'message': "Event does not exist"}, status=404)

        event_doc_id = event_docs[0].id  # Get the document ID
        ticket_cat_ref = db.collection("institutions").document(instId).collection("events").document(event_doc_id).collection("ticket_category").where('category_id','==',categ_id)
        docs = ticket_cat_ref.get()

        if len(docs)>0:
            tc_obj = docs[0].to_dict()
            categ_amount = tc_obj.get('amount', 0)

            for row in worksheet.iter_rows(min_row=2):
                trans_obj = {
                    "transactionNumber": row[0].value,
                    "sender": row[1].value,
                    "senderPhone": row[2].value,
                    "transactionTime": row[3].value,
                    "amount": row[4].value,
                    "referenceNumber": tc_obj['category_id'],
                    "institution_id": instId,
                    'eventId': eventId
                }

                # Add transaction to the transactions subcollection of the event document
                db.collection("institutions").document(instId).collection("events").document(event_doc_id).collection("transactions").add(trans_obj)
                
                if row[4].value is not None:
                    number_of_ticket = int(row[4].value) / int(categ_amount)
                            
                    for i in range(int(number_of_ticket)):
                        ticket_obj = {
                            "name": row[1].value,
                            "phone": str(row[2].value),
                            "amount": 10000,
                            "paymentNumber": str(row[0].value),
                            "paymentTime": datetime.today().strftime('%Y-%m-%d-%H:%M:%S'),
                            "reason": "Buying Ticket",
                            "valid": True,
                            "validatedAt": "",
                            'ticketNumber': str(get_ticket_number()),
                            'singlePurchaseAmount': 10000,
                            "instId": instId,
                            'eventId': eventId
                        }
                        tickets.append(ticket_obj)

            SendSMSAndSaveData_V2(tickets, tickets[0], "EXCEL").start()
            return Response({"tickets": tickets}, status=200)
        else:
            return Response(data={'message': "Ticket category does not exist"}, status=404)

from django.shortcuts import render
    
def get_all_event_ticketCategories(request, institution_id, event_id):
    # method for getting all ticketCategories of event ,and render them to categories page.
    results = db.collection("institutions").document(institution_id).collection("events").document(event_id).collection("ticket_category").get()
    categories = []
    for doc in results:
        ticket_categ_id = doc.id
        ticketCateg_data = doc.to_dict()
        data = {
            'categ_id':ticket_categ_id,
            'name': ticketCateg_data['name'],
            'category_id': ticketCateg_data['category_id'],
            'amount': ticketCateg_data['amount'],
            'group_quantity': ticketCateg_data['group_quantity'],
            'quantity': ticketCateg_data['quantity'],
            'is_free': ticketCateg_data['is_free'],
            'is_group': ticketCateg_data['is_group']
        }
        categories.append(data)
    return render(request, 'categories.html', {'categories': categories, 'institution_id': institution_id, 'event_id': event_id})

def get_event_ticket_category_detail(request, inst_id, event_id,categ_id):
    """
      This method is for getting ticket_category details  of event  and render them to the ticket_category details page .
    """
    response = db.collection("institutions").document(inst_id).collection("events").document(event_id).collection("ticket_category").document(categ_id).get()
    data = None
    if response.exists:
        data=None
        data=response.to_dict()
    return render(request, 'event_ticket_category_details.html', {'category_data': data, 'inst_id': inst_id, 'event_id': event_id,'categ_id':categ_id})
def update_event_ticket_categorise( request,institution_id,event_id,categ_id):
    # this method is for  category update and render to ticketCategory update page
    
    response= db.collection("institutions").document(institution_id).collection("events").document(event_id).collection("ticket_category").document(categ_id).get()
    resp=response.to_dict()
    data={
        'name':resp['name'],
        'category_id':resp['category_id'],
        'amount':resp['amount'],
        'quantity':resp['quantity'],
        'group_quantity':resp['group_quantity'],
        'is_free':resp['is_free'],
        'is_group':resp['is_group'],
    }
    
    return render(request, 'update_event_ticket_categories.html',{'data':data,'institution_id':institution_id,'event_id':event_id,'categ_id':categ_id})
